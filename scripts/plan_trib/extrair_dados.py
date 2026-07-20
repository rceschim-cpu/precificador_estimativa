# -*- coding: utf-8 -*-
"""
Extrai dados tributários da planilha oficial da Controladoria
("Planilha de tributação POSITEC - PLAN_TRIB DD.MM.AAAA.xlsx") e gera
src/planTrib.js com as constantes usadas pelo App:

  ICMS_INTRA  — venda dentro do estado de origem: "NCM|UF" -> {aliq, cred} (%)
  ST_DEST     — ST por destino: "NCM|UFOrigem" -> { UFdest: {mva, aliq} } (%)
                (só destinos onde a planilha diz "SIM"; mva = MVA ORIGINAL)
  DIFAL_INT   — alíquota interna do produto no destino p/ DIFAL:
                "NCM|UFdest" -> aliq (%)  (só quando difere da ALIQ_INT genérica)
  FCP_PROD    — Fundo à Pobreza por produto: "NCM" -> { UF: pct } (%)

Uso:
  python extrair_dados.py "<caminho do xlsx>" "<caminho de saída do .js>"

Colisões (mesmo NCM com variantes diferentes na planilha, ex. Monitor PPB vs
importado): prioriza a variante com "PPB MAO" no nome, depois a com crédito
presumido > 0 (produção incentivada da Positivo), depois a primeira. Todas as
colisões com valores divergentes são listadas em comentário no arquivo gerado.
"""
import openpyxl, sys, json, collections

ALIQ_INT = {"AC":19,"AL":21.5,"AM":20,"AP":18,"BA":20.5,"CE":20,"DF":20,"ES":17,
            "GO":19,"MA":23,"MG":18,"MS":17,"MT":17,"PA":19,"PB":20,"PE":20.5,
            "PI":22.5,"PR":19.5,"RJ":22,"RN":20,"RO":19.5,"RR":20,"RS":17,
            "SC":17,"SE":20,"SP":18,"TO":20}
UFS = list(ALIQ_INT.keys())

def pct(v):
    return round(float(v)*100, 4) if isinstance(v,(int,float)) else None

def score(nome, cred):
    s = 0
    if "PPB MAO" in nome: s += 2
    if cred and cred > 0: s += 1
    return s

def main(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    icms_intra, st_dest, difal_int = {}, collections.defaultdict(dict), {}
    escolhas, colisoes = {}, []

    for uf_dest in UFS:
        ws = wb[f"Destino {uf_dest}"]
        for r in ws.iter_rows(min_row=5, max_row=115, values_only=True):
            ncm, nome, uf_orig = r[1], r[2], r[3]
            if not (isinstance(ncm, str) and "." in ncm and nome and uf_orig):
                continue
            ncm, nome, uf_orig = ncm.strip(), str(nome).strip(), str(uf_orig).strip()
            if uf_orig not in UFS:
                continue
            icms_dest, cred = pct(r[6]), pct(r[7])
            mva, tem_st = pct(r[9]), str(r[10]).strip().upper() if r[10] else ""
            carga_st, carga_difal = pct(r[14]), pct(r[17])
            key = f"{ncm}|{uf_orig}"
            sc = score(nome, cred)

            # resolve colisão de variantes por chave+destino
            ekey = (key, uf_dest)
            if ekey in escolhas:
                prev_nome, prev_sc, prev_vals = escolhas[ekey]
                new_vals = (icms_dest, cred, mva, tem_st, carga_st, carga_difal)
                if new_vals != prev_vals:
                    colisoes.append(f"{key} destino {uf_dest}: '{prev_nome}' (usado) vs '{nome}'")
                if sc <= prev_sc:
                    continue
            escolhas[ekey] = (nome, sc, (icms_dest, cred, mva, tem_st, carga_st, carga_difal))

            # ICMS intra (venda dentro do estado de origem)
            if uf_dest == uf_orig and icms_dest is not None and cred is not None:
                icms_intra[key] = {"aliq": icms_dest, "cred": cred}

            # ST por destino
            if tem_st == "SIM" and mva and carga_st:
                st_dest[key][uf_dest] = {"mva": mva, "aliq": carga_st}
            elif key in st_dest and uf_dest in st_dest[key]:
                del st_dest[key][uf_dest]

            # alíquota interna do produto no destino (p/ DIFAL) — só se difere da genérica
            if carga_difal is not None and abs(carga_difal - ALIQ_INT[uf_dest]) > 0.01:
                difal_int[f"{ncm}|{uf_dest}"] = carga_difal

    # FCP por produto (aba Fundo à Pobreza)
    ws = wb["Fundo à Pobreza"]
    rows = list(ws.iter_rows(min_row=3, max_row=98, values_only=True))
    header = rows[0]
    fcp_ufs = [(j, header[j-1]) for j in range(5, 32) if header[j-1] in ALIQ_INT]
    fcp_prod = collections.defaultdict(dict)
    for r in rows[1:]:
        if not (isinstance(r[1], str) and "." in r[1]):
            continue
        ncm = r[1].strip()
        for j, uf in fcp_ufs:
            v = pct(r[j-1])
            if v:
                prev = fcp_prod[ncm].get(uf)
                if prev is None or v > prev:  # variantes: usa o maior
                    fcp_prod[ncm][uf] = v

    def js(obj):
        return json.dumps(obj, ensure_ascii=False, indent=0, sort_keys=True)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("// GERADO por scripts/plan_trib/extrair_dados.py — NÃO editar na mão.\n")
        f.write("// Fonte: Planilha de tributação POSITEC (PLAN_TRIB) da Controladoria.\n")
        f.write("// Regerar sempre que a Controladoria publicar planilha nova.\n")
        if colisoes:
            f.write("// COLISÕES de variantes resolvidas automaticamente (revisar se necessário):\n")
            for c in sorted(set(colisoes)):
                f.write(f"//   {c}\n")
        f.write("\n// Venda DENTRO do estado de origem: \"NCM|UF\" -> {aliq, cred} em %\n")
        f.write(f"export const ICMS_INTRA = {js(icms_intra)};\n")
        f.write("\n// ST por destino: \"NCM|UFOrigem\" -> { UFdest: {mva, aliq} } em % (mva ORIGINAL)\n")
        f.write(f"export const ST_DEST = {js({k:v for k,v in st_dest.items() if v})};\n")
        f.write("\n// Alíquota interna do produto no destino p/ DIFAL (só quando difere da genérica): \"NCM|UFdest\" -> %\n")
        f.write(f"export const DIFAL_INT = {js(difal_int)};\n")
        f.write("\n// Fundo à Pobreza por produto: \"NCM\" -> {UF: pct}\n")
        f.write(f"export const FCP_PROD = {js(dict(fcp_prod))};\n")

    print(f"ICMS_INTRA: {len(icms_intra)} chaves")
    print(f"ST_DEST: {len([k for k,v in st_dest.items() if v])} chaves NCM|origem")
    print(f"DIFAL_INT: {len(difal_int)} chaves")
    print(f"FCP_PROD: {len(fcp_prod)} NCMs")
    print(f"Colisões: {len(set(colisoes))}")
    print(f"Gerado: {out_path}")

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
