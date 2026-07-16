# -*- coding: utf-8 -*-
import openpyxl, json, sys, datetime, collections

MES_PT = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}

def load_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Base Precificado"]
    headers = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    idx = {h:i for i,h in enumerate(headers) if h}
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[idx['Valor Ou Ajuste']] == "Valor" and row[idx['Venda ou Devolução']] == "Venda":
            rows.append(row)
    return rows, idx

def sq(v):
    # sql-safe string
    if v is None: return "NULL"
    if isinstance(v,(int,float)): return str(v)
    return "'" + str(v).replace("'","''") + "'"

def build_indices(path, mes_ano_label):
    rows, idx = load_rows(path)
    groups = collections.defaultdict(list)
    for r in rows:
        key = (r[idx['Material']], r[idx['Cliente']], r[idx['Canal']])
        groups[key].append(r)

    inserts = []
    for key, rs in groups.items():
        rs.sort(key=lambda r: r[idx['Data doc fatur']] or datetime.datetime.min)
        r = rs[-1]  # mais recente
        def m(col):
            v = r[idx[col]]
            return abs(v) if v is not None else 0
        sku, cliente, canal = key
        cliente_nome = r[idx['Nome']]
        planta = r[idx['Planta']]
        data_ref = r[idx['Data doc fatur']]
        data_ref_s = data_ref.strftime("%Y-%m-%d") if hasattr(data_ref,"strftime") else data_ref
        row = {
            "sku": str(sku), "canal": str(canal), "cliente": str(cliente), "cliente_nome": cliente_nome,
            "planta": planta or "", "data_ref": data_ref_s,
            "ipi_pct": m('IPVA - IPI Taxa de imposto - Montante_') if 'IPVA - IPI Taxa de imposto - Montante_' in idx else 0,
            "icms_pct": m('ICVA - ICMS Taxa de imposto - Montante_'),
            "cred_pct": m('ZI02 - Crédito Presum ICMS - Montante_') if 'ZI02 - Crédito Presum ICMS - Montante_' in idx else 0,
            "fti_pct": m('ZFTI - Taxa FTI (%) - Montante') if 'ZFTI - Taxa FTI (%) - Montante' in idx else 0,
            "pd_pct": m('ZV25 - P&D (%) - Montante_'),
            "scrap_pct": m('ZV29 - Custo Quebr+Scrap(%) - Montante_'),
            "frete_pct": m('ZV07 - Frete Saída (%) - Montante_') if 'ZV07 - Frete Saída (%) - Montante_' in idx else 0,
            "zv09_pct": m('ZV09 - Custo Financeiro (%) - Montante_'),
            "zv11_pct": m('ZV11 - CustoFixo-Precif(%) - Montante_'),
            "mkt_pct": m('ZV12 - Marketing-Precif(%) - Montante_'),
            "bkp_pct": m('ZV23 - BKP Peças (%) - Montante_'),
            "rebate_pct": m('ZV28 - Rebate - Montante_'),
            "margger_pct": m('ZV05 - ÍndGerencial-Precif - Montante_'),
            "margem_pct": (r[idx['ZM01 - Margem (%) - Montante_']] or 0),
            "mc_pct": (r[idx['MC - %']] or 0),
            "ml_pct": (r[idx['ML - %']] or 0),
            "fonte": mes_ano_label,
        }
        inserts.append(row)
    return inserts

if __name__ == "__main__":
    path = sys.argv[1]
    label = sys.argv[2]
    out = sys.argv[3]
    rows = build_indices(path, label)
    print(f"{path}: {len(rows)} linhas de precificacao_indices geradas")
    with open(out, "w", encoding="utf-8") as f:
        cols = ["sku","canal","cliente","cliente_nome","planta","data_ref","ipi_pct","icms_pct","cred_pct","fti_pct",
                "pd_pct","scrap_pct","frete_pct","zv09_pct","zv11_pct","mkt_pct","bkp_pct","rebate_pct",
                "margger_pct","margem_pct","mc_pct","ml_pct","fonte"]
        f.write(f"insert into precificacao_indices ({','.join(cols)}) values\n")
        vals = []
        for row in rows:
            vals.append("(" + ",".join(sq(row[c]) for c in cols) + ")")
        f.write(",\n".join(vals))
        f.write(";\n")
