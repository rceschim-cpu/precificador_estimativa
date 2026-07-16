# -*- coding: utf-8 -*-
import openpyxl, json, sys, collections

def load_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Base Precificado"]
    headers = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    idx = {h:i for i,h in enumerate(headers) if h}
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[idx['Valor Ou Ajuste']] == "Valor" and row[idx['Venda ou Devolução']] == "Venda":
            rows.append(row)
    return rows, idx

def sq(v):
    if v is None: return "NULL"
    if isinstance(v,(int,float)): return str(v)
    return "'" + str(v).replace("'","''") + "'"

def build_custos(path, mes_ano_label, data_ref_mes, taxas_ref):
    rows, idx = load_rows(path)
    groups = collections.defaultdict(list)
    for r in rows:
        key = (r[idx['Material']], r[idx['Planta']])
        groups[key].append(r)

    def col(name):
        return idx.get(name)

    i_qtd = idx['Qtd faturada efet']
    i_zt03 = idx['ZT03 - Total Preço - Valor_']
    i_cmv = idx['CMV']
    i_gar = idx['GARANTIA']
    i_backup = idx['Backup']
    i_difal = idx.get('ZV03 - DIFAL Valor')
    i_st = idx.get('ZSTI - ICMS ST s/ Faturam - Valor_')
    i_ncm = idx['NCM']
    i_custousd_v = idx.get('ZV01 - Custo PrdDolar(Vlr) - Valor_')
    i_taxadolar_m = idx.get('ZV15 - Taxa Dolar - Precf - Montante_')
    i_custotransf_v = idx.get('ZV14 - Custo Transfor (Vlr) - Valor_')
    i_ggf = idx['GGF']
    i_mcrs = idx['MC - R$']
    i_mlrs = idx['ML - R$']

    inserts = []
    skipped_sem_taxas = []
    for (sku, planta), rs in groups.items():
        vol = sum((r[i_qtd] or 0) for r in rs)
        if not vol: continue
        receita = sum((r[i_zt03] or 0) for r in rs)
        cmv_unit = abs(sum((r[i_cmv] or 0) for r in rs)) / vol
        garantia_pct = (abs(sum((r[i_gar] or 0) for r in rs)) / receita * 100) if receita else 0
        backup_pct = (abs(sum((r[i_backup] or 0) for r in rs)) / receita * 100) if receita else 0
        difal_pct = (abs(sum((r[i_difal] or 0) for r in rs)) / receita * 100) if (receita and i_difal is not None) else 0
        st_pct = (abs(sum((r[i_st] or 0) for r in rs)) / receita * 100) if (receita and i_st is not None) else 0
        custo_usd_total = abs(sum((r[i_custousd_v] or 0) for r in rs)) if i_custousd_v is not None else 0
        custo_usd_unit = custo_usd_total / vol
        taxa_dolar_vals = [abs(r[i_taxadolar_m]) for r in rs if i_taxadolar_m is not None and r[i_taxadolar_m]]
        taxa_dolar = sum(taxa_dolar_vals)/len(taxa_dolar_vals) if taxa_dolar_vals else 0
        custo_transf_unit = (abs(sum((r[i_custotransf_v] or 0) for r in rs)) / vol) if i_custotransf_v is not None else 0
        ggf_unit = abs(sum((r[i_ggf] or 0) for r in rs)) / vol
        ncm = rs[0][i_ncm]
        preco_medio = receita/vol if vol else 0
        mc_pct = (sum((r[i_mcrs] or 0) for r in rs) / receita) if receita else 0
        ml_pct = (sum((r[i_mlrs] or 0) for r in rs) / receita) if receita else 0

        tax = taxas_ref.get((str(sku), planta))
        if tax is None:
            skipped_sem_taxas.append((sku, planta))
            icms_pct = ipi_pct = cred_presum_pct = fti_pct = 0
        else:
            icms_pct, ipi_pct, cred_presum_pct, fti_pct = tax

        row = {
            "sku": str(sku), "planta": planta or "", "ncm": ncm, "data_ref": data_ref_mes,
            "volume": vol, "receita_liq": round(receita,2), "preco_medio": round(preco_medio,2),
            "custo_usd_unit": round(custo_usd_unit,2), "taxa_dolar": round(taxa_dolar,4),
            "custo_brl_unit": 0, "custo_transf_unit": round(custo_transf_unit,2),
            "ggf_unit": round(ggf_unit,2), "cmv_unit": round(cmv_unit,2),
            "garantia_pct": round(garantia_pct,4), "backup_pct": round(backup_pct,4),
            "st_pct": round(st_pct,4), "difal_pct": round(difal_pct,4),
            "icms_pct": icms_pct, "ipi_pct": ipi_pct, "cred_presum_pct": cred_presum_pct, "fti_pct": fti_pct,
            "mc_pct": round(mc_pct,4), "ml_pct": round(ml_pct,4),
            "fonte": mes_ano_label,
        }
        inserts.append(row)
    return inserts, skipped_sem_taxas

if __name__ == "__main__":
    path, label, data_ref_mes, taxas_json, out = sys.argv[1:6]
    taxas_raw = json.load(open(taxas_json, encoding="utf-8"))
    taxas_ref = {}
    for r in taxas_raw:
        key = (str(r['sku']), r['planta'])
        if key not in taxas_ref:
            taxas_ref[key] = (r['icms_pct'] or 0, r['ipi_pct'] or 0, r['cred_presum_pct'] or 0, r['fti_pct'] or 0)

    rows, skipped = build_custos(path, label, data_ref_mes, taxas_ref)
    print(f"{path}: {len(rows)} linhas de precificacao_custos geradas; {len(skipped)} sem referencia de impostos (ficaram 0)")

    cols = ["sku","planta","ncm","data_ref","volume","receita_liq","preco_medio","custo_usd_unit","taxa_dolar",
            "custo_brl_unit","custo_transf_unit","ggf_unit","cmv_unit","garantia_pct","backup_pct","st_pct",
            "difal_pct","icms_pct","ipi_pct","cred_presum_pct","fti_pct","mc_pct","ml_pct","fonte"]
    with open(out, "w", encoding="utf-8") as f:
        f.write(f"insert into precificacao_custos ({','.join(cols)}) values\n")
        vals = []
        for row in rows:
            vals.append("(" + ",".join(sq(row[c]) for c in cols) + ")")
        f.write(",\n".join(vals))
        f.write(";\n")
